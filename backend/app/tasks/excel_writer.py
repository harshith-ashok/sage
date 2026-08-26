"""MS Office support, part 1: a real .xlsx spreadsheet writer (openpyxl) —
named in this project's Stack from the start ("python-docx/pptx/openpyxl")
but only the docx side had been built until now. Tabular data (an ML fit's
inputs, a calculation's results, anything the user asks to see "as Excel"
rather than crammed into a Word table) belongs in a real spreadsheet: real
numeric cells (sortable, usable in a formula), not a table image or a wall
of text.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def write_excel(path: str, sheet_name: str, headers: list[str], rows: list[list]) -> None:
    """Writes one sheet: a bold header row plus `rows` of data. Cells that
    are already int/float are written as real numbers (not stringified) so
    the result is usable in Excel formulas/sorting, not just readable."""
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Sheet1")[:31]  # Excel's own sheet-name length limit

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row[: len(headers)], start=1):
            ws.cell(row=r, column=c, value=value)

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        values = [str(headers[col - 1])] + [str(row[col - 1]) if col - 1 < len(row) else "" for row in rows]
        ws.column_dimensions[letter].width = min(max(len(v) for v in values) + 2, 40)

    ws.freeze_panes = "A2"
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
